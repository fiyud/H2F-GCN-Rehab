import os
import csv
import openpyxl
import numpy as np
import pandas as pd

def load_kimore_data(path="Kimore", enable_kinect_joints=True, enable_slice_list=False):
    kinect_joints = ["spinebase", "spinemid", "neck", "head", 
                     "shoulderleft", "elbowleft", "wristleft", 
                     "handleft", "shoulderright", "elbowright", 
                     "wristright", "handright", "hipleft", "kneeleft", 
                     "ankleleft", "footleft", "hipright", "kneeright", 
                     "ankleright", "footright", "spineshoulder", "handtipleft", 
                     "thumbleft", "handtipright", "thumbright"]

    data = []
    for (root, dirs, files) in os.walk(path):
        if "Raw" in dirs:
            new_dict = {}
            
            # get exercise number
            new_dict["Exercise"] = int(root[-1])

            # extract raw data
            raw_files = os.listdir(os.path.join(root, "Raw"))
            for file in raw_files:
                file_path = os.path.join(os.path.join(root, "Raw"), file)
                csv_file = open(file_path, newline='')
                
                csv_reader = csv.reader(csv_file)
               
                if file.startswith("JointOrientation"):
                    if enable_kinect_joints:
                        for joint in kinect_joints:
                            new_dict[joint + "-o"] = []

                        for row in csv_reader:
                            for i in range(len(kinect_joints)):
                                if len(row) > 0:
                                    new_dict[kinect_joints[i] + "-o"].append(np.array([float(i) for i in row[(4*i):(4*i+4)]]))
                    else:
                        new_dict["JointOrientation"] = []
                        for row in csv_reader:
                            if len(new_dict["JointOrientation"]) >= 182 and enable_slice_list:
                                break
                            elif len(row) > 0:
                                if '' in row:
                                    row.remove('')
                                new_dict["JointOrientation"].append(np.array([float(i) for i in row]))

                elif file.startswith("JointPosition"):
                    if enable_kinect_joints:
                        for joint in kinect_joints:
                            new_dict[joint + "-p"] = []

                        for row in csv_reader:
                            for i in range(len(kinect_joints)):
                                if len(row) > 0:
                                    new_dict[kinect_joints[i] + "-p"].append(np.array([float(i) for i in row[(4*i):(4*i+3)]]))
                    else:
                        new_dict["JointPosition"] = []
                        for row in csv_reader:
                            if len(new_dict["JointPosition"]) >= 182 and enable_slice_list:
                                break
                            elif len(row) > 0:
                                if '' in row:
                                    row.remove('')
                                new_dict["JointPosition"].append(np.array([float(i) for i in row]))

                elif file.startswith("TimeStamp"):
                    new_dict["Timestamps"] = []
                    for row in csv_reader:
                        if len(new_dict["Timestamps"]) >= 182 and enable_slice_list:
                            break
                        elif len(row) > 0:
                            if '' in row:
                                    row.remove('')
                            new_dict["Timestamps"].append(row)

            # extract data labels
            label_files = os.listdir(os.path.join(root, "Label"))
            for file in label_files:
                file_path = os.path.join(os.path.join(root, "Label"), file)
                book = openpyxl.load_workbook(file_path)
                sheet = book.active

                if file.startswith("SuppInfo"):
                    for i in range(1, sheet.max_column):
                        t = sheet.cell(1, i).value
                        v = sheet.cell(2, i).value
                        new_dict[t] = v
                
                elif file.startswith("ClinicalAssessment"):
                    new_dict["cTS"] = sheet.cell(2, new_dict["Exercise"]+1).value
                    new_dict["cPO"] = sheet.cell(2, new_dict["Exercise"]+6).value
                    new_dict["cCF"] = sheet.cell(2, new_dict["Exercise"]+11).value
            
            data.append(new_dict)
            
    return data

def preprocess_merged_data(df):
    columns_o = [col for col in df.columns if col.endswith('-o')]
    columns_p = [col.replace('-o', '-p') for col in columns_o]

    df_merged = pd.DataFrame()

    for col_o, col_p in zip(columns_o, columns_p):
        feature_name = col_o.replace('-o', '') 
        merged_column = [np.hstack((o, p)) for o, p in zip(df[col_o], df[col_p])]
        df_merged[feature_name] = merged_column
    
    df_merged['cTS'] = df['cTS'] / df['cTS'].max()
    
    return df_merged